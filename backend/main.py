import shutil
import networkx as nx
import pandas as pd
import matplotlib.pyplot as plt
from fastapi import FastAPI, File, UploadFile
from starlette.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
import uvicorn
import csv
import base64
import os
#import ssl

app = FastAPI() # Instancia de FastAPI

#definimos el contexto ssl para exponer la api a traves de https
#ssl_context = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
#ssl_context.load_cert_chain('/home/bitnami/cert.pem', keyfile='/home/bitnami/key.pem')
#ssl_context.load_cert_chain('./cert.pem', keyfile='./key.pem')

#origenes permitidos por el CORS
origins = [
  "https://graph-front-ksqw.vercel.app",  # El dominio de tu frontend en Vercel
  "http://localhost",                    # Para desarrollo local (si tu frontend corre en localhost)
  "http://localhost:8000",               # Otros puertos comunes para desarrollo local
  "http://localhost:4200",               # Puerto común para Angular en desarrollo
  "http://localhost:3000",               # Puerto común para React/Next.js en desarrollo
  "http://localhost:5000",
]

#añadimos el middelware (CORS)
app.add_middleware(
  CORSMiddleware,
  allow_origins=origins,
  allow_credentials=True,
  allow_methods=["*"],
  allow_headers=["*"],
)

#Funcion que genera los datos en csv para el grafo global
def crear_datos_grafo_global(fichero, diametroF: int, diametroK: int, multiplicador_algoritmo: float, multiplicador_pesos: float, tamano_texto_nodos: int, opacidadNodos: float, habilidad: bool):
    df = pd.read_excel(fichero + '.xlsx', skiprows=1)    #Cargamos el archivo excel con los datos de entrada
    suma_pesos_totales = 0  #inicializamos la suma de los pesos totales
    pesos_totales = []  #generamos un array de pesos totales
 
    #escribimos los datos en el archivo CSV
    with open(fichero + '.csv', 'w', newline='') as archivo_csv:
        datos_csv = csv.writer(archivo_csv)  #instancia de escritura de CSV
        #Escribimos los encabezados del CSV
        datos_csv.writerow(['Nivel aprendizaje Taxonomia de Bloom', 'Ra', 'Descripcion', 'Steam', 'Peso'])
        #Iteramos sobre las filas del DataFrame
        for index, row in df.iterrows():
            #Revisamos si la primera columna es NaN. Si lo es, se lo salta
            if pd.isna(row.iloc[0]) and pd.isna(row.iloc[3]):
                continue
            elif pd.isna(row.iloc[0]) and pd.notna(row.iloc[3]):    #Si la primera columna es NaN y la cuarta no...
                #...iteramos sobre las columnas numéricas (STEAMH)...
                for col in df.columns[3:9]:
                    suma_pesos_totales += row[col]  #...sumamos los pesos en la variable de pesos totales...
                    col_mostrada = col  #definimos la columna con la letra a mostrar
                    #...si la habilidad es 'K' y hemos cambiado la letra por 'P' en el frontend...
                    if col == 'K' and habilidad:
                        col_mostrada = 'P'  #...ponemos la columna mostrada como 'P'
                    #...añadimos el peso de cada letra al array de pesos totales...
                    pesos_totales.append(
                        {
                            "letra": col_mostrada,
                            "valor": row[col], 
                        }    
                    )
                break   #...y salimos del bucle
            else: #...en caso contrario...          
                #...iteramos sobre las columnas numéricas...
                for col in df.columns[3:10]:
                    #...si la columna es NaN, se lo salta...
                    if pd.isna(row[col]) or row[col] == 0 or row[col] == '0':
                        continue   #...y salimos del bucle
                    
                    col_mostrada = col  #definimos la columna con la letra a mostrar
                    #...si la habilidad es 'K' y hemos cambiado la letra por 'P' en el frontend...
                    if col == 'K' and habilidad:
                        col_mostrada = 'P'  #...ponemos la columna mostrada como 'P'
                    #Escribimos una línea en el CSV por cada columna numérica
                    datos_csv.writerow(list(row[:3]) + [col_mostrada] + [row[col]])

    #Cargamos el csv con los datos de entrada
    datos_entrada = pd.read_csv(fichero + '.csv', encoding='ISO-8859-1')
    #intentamos crear el grafo, si se crea correctamente devolvemos 'true', 'false' en caso contrario
    try:
        #llamamos a la funcion que genera el grafo global
        crear_grafo_global(datos_entrada, diametroF, diametroK, multiplicador_algoritmo, multiplicador_pesos, suma_pesos_totales, pesos_totales, tamano_texto_nodos, opacidadNodos, habilidad, fichero)
        return True
    except:
        return False

#Funcion que genera los datos en csv para el grafo ODS
def crear_datos_grafo_ods(tamano_texto_nodos, opacidadNodos, fichero):
    workbook = load_workbook(fichero + '.xlsx') #cargamos el fichero xlsx y guardamos el contenido en la variable
    hoja = workbook.active  #capturamos el nombre de la hoja activa del fichero xlsx leido
    
    #escribimos los datos en el archivo CSV
    with open(fichero + '.csv', 'w', newline='') as archivo_csv:
        #instancia de escritura de CSV
        datos_csv = csv.writer(archivo_csv)
        #Escribimos los encabezados del CSV
        datos_csv.writerow(['Competencia', 'ODS'])
    
        #para cada fila del fichero xlsx leido
        for fila in hoja.iter_rows(values_only=True):
            #si la primera fila del fichero (fila de cabecera) son numeros, nos hemos equivocado de fichero, devolvemos false
            if(type(fila[1]) == str):
                return False
            
            #si la primera celda está vacia, querra decir que es la fila de titulos de cabecera y por tanto la descartamos...
            if(fila[0] == None):
                continue
            
            competencia = fila[0]   #...y asignamos a la variable de competencia la primera celda de las filas evaluada
            #para cada celda de la fila que estamos evaluando...
            for i in range(len(fila)):
                #...si es una 'X' insertamos en el fichero csv la competencia y su numero de ods (titulo de columna ods)...           
                if(str(fila[i]).upper() == 'X'):
                    datos_csv.writerow([competencia, i])
                elif(fila[i] == None):  #...y si está vacia, la descartamos
                    continue
    
    #Cargamos el csv con los datos de entrada
    datos_entrada = pd.read_csv(fichero + '.csv', encoding='ISO-8859-1')
    
    #intentamos crear el grafo, si se crea correctamente devolvemos 'true', 'false' en caso contrario
    try:
        #llamamos a la funcion que genera el grafo
        crear_grafo_ods(datos_entrada, tamano_texto_nodos, opacidadNodos, fichero)
        return True
    except:
        return False

#Funcion que genera los datos en csv para el grafo STEAM
def crear_datos_grafo_steam(tamano_texto_nodos, opacidadNodos, fichero):
    workbook = load_workbook(fichero + '.xlsx') #cargamos el fichero xlsx y guardamos el contenido en la variable
    hoja = workbook.active  #capturamos el nombre de la hoja activa del fichero xlsx leido
    
    #escribimos los datos en el archivo CSV
    with open(fichero + '.csv', 'w', newline='') as archivo_csv:
        #instancia de escritura de CSV
        datos_csv = csv.writer(archivo_csv)
        #Escribimos los encabezados del CSV
        datos_csv.writerow(['STEAM', 'Conexion'])
    
        #para cada fila del fichero xlsx leido
        for fila in hoja.iter_rows(values_only=True):             
            #si la primera fila del fichero (fila de cabecera) son numeros, nos hemos equivocado de fichero, devolvemos false
            if(type(fila[1]) == int):
                return False
            
            #si en alguna fila del fichero aparece la palabra 'RA', nos hemos equivocado de fichero, devolvemos false
            if(fila[1] == 'RA'):
                return False
                      
            #si la primera celda está vacia, querra decir que es la fila de titulos de cabecera y por tanto la descartamos...
            if(fila[0] == None):
                continue
            
            steam = fila[0]   #...y asignamos a la variable steam la primera celda de las filas evaluada (letra steam)
            #para cada celda de la fila que estamos evaluando...
            for i in range(len(fila)):
                #...si es una 'X' insertamos en el fichero csv la letra steam y su numero de letra steam...           
                if(str(fila[i]).upper() == 'X'):
                    datos_csv.writerow([steam, steam[0:1] + str(i)])
                elif(fila[i] == None):  #...y si está vacia, la descartamos
                    continue
    
    #Cargamos el csv con los datos de entrada
    datos_entrada = pd.read_csv(fichero + '.csv', encoding='ISO-8859-1')
    
    #intentamos crear el grafo, si se crea correctamente devolvemos 'true', 'false' en caso contrario
    try:
        #llamamos a la funcion que genera el grafo
        crear_grafo_steam(datos_entrada, tamano_texto_nodos, opacidadNodos, fichero)
        return True
    except:
        return False

#funcion que genera el grafo global en funcion de los datos de entrada en fichero csv
def crear_grafo_global(datos, diametroF, diametroK, multiplicador_algoritmo, multiplicador_pesos, suma_pesos_totales, pesos_totales, tamano_texto_nodos, opacidadNodos, habilidad, nombre_fichero):
    #Generamos una instancia del grafo
    G = nx.Graph()

    k_nodes = []    #generamos un array de nodos que conectaran con la disciplina 'k', ya que sus conexiones seran de color marron
    edge_colors = []    #generamos un array de colores para las aristas

    #Añadimos aristas al grafo con pesos
    for idx, row in datos.iterrows():
        G.add_edge(row['Ra'], row['Steam'], weight=row['Peso'])
        #Si el nodo conecta con la disciplina 'K' o 'P' lo añadimos al array de nodos que conectan con 'K' o 'P'
        if (row['Steam'] == 'K' or row['Steam'] == 'P'):
            k_nodes.append(row['Ra'])
            
    #Añadimos aristas al array de colores de aristas
    for n2 in G.edges():
        #Si la arista en cualquiera de las dos posiciones de G.edges() está en el array de nodos que conectan con la disciplina 'K' o 'P',
        #la añadimos a edge_colors con color marron, si no, con color gris
        if n2[0] in k_nodes or n2[1] in k_nodes:
            edge_colors.append('brown')
        else:
            edge_colors.append('gainsboro')

    #Calculamos la centralidad de intermediación (Betweenness Centrality) para cada nodo (es el tipo de grafo que estamos utilizando)
    betweenness_centrality = nx.betweenness_centrality(G)

    node_colors = []    #generamos un array de colores para los nodos
    node_sizes = [] #generamos un array de tamaños para los nodos
    
    #Asignamos un color a cada nodo de disciplinas STEAMHK y gris para el resto
    for n in G.nodes():
        if n == 'S':
            node_colors.append('green')            
        elif n == 'T':
            node_colors.append('blue')
        elif n == 'E':
            node_colors.append('orange')
        elif n == 'A':
            node_colors.append('violet')
        elif n == 'M':
            node_colors.append('red')
        elif n == 'H':
            node_colors.append('cyan')
        elif n == 'K' or n == 'P':
            node_colors.append('brown')
        else:
            node_colors.append('grey')

    #para cada nodo asignamos un tamaño proporcional a su centralidad de intermediación y su peso en la excel
    for key, value in betweenness_centrality.items():
        #Si el nodo no es una disciplina STEAMHK (tiene mas de una letra) le asignamos el tamaño que viene desde el frontend
        #y si no, el tamaño será proporcional a su centralidad de intermediación y a su peso en la excel
        if len(key) > 1:
            node_sizes.append(diametroF * int(tamano_texto_nodos) / 5)    #el diametro de los nodos RA es el que se define en la llamada al endpoint
        else:
            #Si el nodo es la disciplina 'K' le asignamos un tamaño fijo
            if key == 'K' or key == 'P':
                node_sizes.append(diametroK * int(tamano_texto_nodos) / 5)
            else: 
                matches = [e for e in pesos_totales if e['letra']  == key][0]
                node_sizes.append(((multiplicador_algoritmo * value * 2 * int(tamano_texto_nodos)) + 1) * (((multiplicador_pesos * matches['valor']) / suma_pesos_totales * 70000) + 1))

    #Definimos el area que contendrá el grafo
    fig, ax = plt.subplots(figsize=(10, 8))

    #Definimos el titulo del grafo
    ax.set_title('Global Graph')

    #Calculamos la posición de los nodos en el grafico
    pos = nx.shell_layout(G, scale=.1)
    
    #En funcion de si es 'K' o 'P' definimos la posicion fija del nodo 'K' para que no se salga del grafo
    if habilidad:
        pos['P'][1] = 0.01
        pos['P'][0] = 0.01
    else:
        pos['K'][1] = 0.01
        pos['K'][0] = 0.01

    #Dibujamos el grafo
    nx.draw_networkx(
        G,  #datos completos del grafo
        pos=pos,    #posicion de los nodos
        with_labels=True,   #etiquetas visibles (nombres de los nodos)
        node_color=node_colors,  #color de los nodos
        node_size=node_sizes,    #tamaño de los nodos
        alpha=opacidadNodos,  #opacidad de los nodos
        edge_color=edge_colors, #color de las aristas
        width=3,  #grosor de las aristas
        font_size=tamano_texto_nodos,   #tamaño del texto de los nodos
        font_weight='bold' #ponemos los textos de los nodos en negrita
    )
    
    plt.axis('on')  #Activamos los ejes (bordes del grafico)
    plt.tight_layout()  #Ajustamos el grafico para que no se salga del area de visualizacion
    plt.savefig(nombre_fichero + '.png', bbox_inches='tight', pad_inches=0.1)   #Guardamos el grafico
    plt.close() #Cerramos el grafico
    
#funcion que genera el grafo ODS en funcion de los datos de entrada en fichero csv
def crear_grafo_ods(datos, tamano_texto_nodos, opacidadNodos, nombre_fichero):
    #Generamos una instancia del grafo
    G = nx.Graph()

    #Añadimos aristas al grafo con pesos
    for idx, row in datos.iterrows():
        G.add_edge(row['Competencia'], row['ODS'])

    #Calculamos la centralidad de intermediación (Betweenness Centrality) para cada nodo (es el tipo de grafo que estamos utilizando)
    betweenness_centrality = nx.betweenness_centrality(G)

    edge_colors = []    #generamos un array de colores para las aristas
    node_colors = []    #generamos un array de colores para los nodos
    node_sizes = [] #generamos un array de tamaños para los nodos
    
    #Asignamos un color a cada nodo de disciplinas STEAMHK y gris para el resto
    for n in G.nodes():
        if n == 1:
            node_colors.append('#E32444')            
        elif n == 2:
            node_colors.append('#DFA638')
        elif n == 3:
            node_colors.append('#4BA246') 
        elif n == 4:
            node_colors.append('#C82130')
        elif n == 5:
            node_colors.append('#EC442C')
        elif n == 6:
            node_colors.append('#21BEE6')
        elif n == 7:
            node_colors.append('#F9C420')
        elif n == 8:
            node_colors.append('#A41D46')            
        elif n == 9:
            node_colors.append('#F26B33')
        elif n == 10:
            node_colors.append('#DE1D69')
        elif n == 11:
            node_colors.append('#F99E2A')
        elif n == 12:
            node_colors.append('#BF8D2C')
        elif n == 13:
            node_colors.append('#3D8147')
        elif n == 14:
            node_colors.append('#1F96D2')
        elif n == 15:
            node_colors.append('#58B949')
        elif n == 16:
            node_colors.append('#056CA2')
        elif n == 17:
            node_colors.append('#0D4A6C')
        else:
            node_colors.append('grey')

    #para cada nodo calculado por el algoritmo, asignamos el tamaño y el color
    for key in betweenness_centrality.items():
        if(str(key[0]).isdigit()):
            if(key[1] == 0):
                node_sizes.append(100 * int(tamano_texto_nodos))
            else:
                node_sizes.append(key[1] * 30000 * int(tamano_texto_nodos))
        else:
            node_sizes.append(200 * int(tamano_texto_nodos))
            
        edge_colors.append('grey')                

    #Definimos el area que contendrá el grafo
    fig, ax = plt.subplots(figsize=(10, 8))

    #Definimos el titulo del grafo
    ax.set_title('ODS Graph')
    
    #Calculamos la posición de los nodos en el grafico
    pos = nx.shell_layout(G, scale=2)

    #Dibujamos el grafo
    nc = nx.draw_networkx(
        G,  #datos completos del grafo
        pos=pos,    #posicion de los nodos
        with_labels=True,   #etiquetas visibles (nombres de los nodos)
        node_color=node_colors,  #color de los nodos
        node_size=node_sizes,    #tamaño de los nodos
        alpha=opacidadNodos,  #opacidad de los nodos
        edge_color=edge_colors, #color de las aristas
        width=3,  #grosor de las aristas
        font_size=tamano_texto_nodos,   #tamaño del texto de los nodos
        font_weight='bold', #ponemos los textos de los nodos en negrita
        label=''    #texto de la leyenda (lo dejamos vacio para que el grafo no se salga de los limites)
    )

    plt.axis('on')  #Activamos los ejes (bordes del grafico)
    plt.tight_layout()  #Ajustamos el grafico para que no se salga del area de visualizacion
    plt.savefig(nombre_fichero + '.png', bbox_inches='tight', pad_inches=0.1)   #Guardamos el grafico
    plt.close() #Cerramos el grafico
    
#funcion que genera el grafo STEAM en funcion de los datos de entrada en fichero csv
def crear_grafo_steam(datos, tamano_texto_nodos, opacidadNodos, nombre_fichero):
    #Generamos una instancia del grafo
    G = nx.Graph()

    edge_colors = []    #generamos un array de colores para las aristas

    #Añadimos aristas al grafo con pesos
    for idx, row in datos.iterrows():
        G.add_edge(row['STEAM'], row['Conexion'])

    primera_columna = datos.iloc[:, 0]  #nos quedamos con la primera columna de los datos de entrada
    #agrupamos por el numero de apariciones para ver cuantas veces aparece (tiene conexion) cada nodo
    count = pd.Series(primera_columna).value_counts()

    node_colors = []    #generamos un array de colores para los nodos
    node_sizes = [] #generamos un array de tamaños para los nodos
    
    #Asignamos un color a cada nodo de disciplinas STEAMHK y gris para el resto
    for n in G.nodes():
        if n[0:1] == 'S':
            node_colors.append('green')            
        elif n[0:1] == 'T':
            node_colors.append('blue')
        elif n[0:1] == 'E':
            node_colors.append('orange')
        elif n[0:1] == 'A':
            node_colors.append('violet')
        elif n[0:1] == 'M':
            node_colors.append('red')
        else:
            node_colors.append('grey')
            
        #para cada nodo dentro de las agrupaciones...
        for x in range(0, len(count)):
            #...si el nodo grafo es el mismo que el de las agrupaciones...
            if(n == count.keys()[x]):
                #...cogemos el numero de apariciones de ese nodo y metemos ese valor
                #multiplicado por 500 y por el tamaño del nodo definido en el array de tamaños de nodo
                node_sizes.append(500 * count.iloc[x] * int(tamano_texto_nodos))
        
        edge_colors.append('grey')   #ponemos todas las aristas en gris

    #Definimos el area que contendrá el grafo
    fig, ax = plt.subplots(figsize=(12, 10))

    #Definimos el titulo del grafo
    ax.set_title('STEAM Graph')

    #Calculamos la posición de los nodos en el grafico
    pos = nx.shell_layout(G, scale=.1)

    #Dibujamos el grafo
    nc = nx.draw_networkx(
        G,  #datos completos del grafo
        pos=pos,    #posicion de los nodos
        with_labels=True,   #etiquetas visibles (nombres de los nodos)
        node_color=node_colors,  #color de los nodos
        node_size=node_sizes,    #tamaño de los nodos
        alpha=opacidadNodos,  #opacidad de los nodos
        edge_color=edge_colors, #color de las aristas
        width=3,  #grosor de las aristas
        font_size=tamano_texto_nodos,   #tamaño del texto de los nodos
        font_weight='bold', #ponemos los textos de los nodos en negrita
        label=''    #texto de la leyenda (lo dejamos vacio para que el grafo no se salga de los limites)
    )

    plt.axis('on')  #Activamos los ejes (bordes del grafico)
    plt.tight_layout()  #Ajustamos el grafico para que no se salga del area de visualizacion
    plt.savefig(nombre_fichero + '.png', bbox_inches='tight', pad_inches=0.1)   #Guardamos el grafico
    plt.close() #Cerramos el grafico

#funcion que borra los ficheros temporales usados para la generacion del grafo
def borrar_ficheros_temporales(ficheroXLSX, ficheroCSV, ficheroPNG):
    #borramos el fichero xlsx temporal
    try:
        os.remove(ficheroXLSX)
    except FileNotFoundError:
        print("El archivo " + ficheroXLSX + " no se encontró")
    except PermissionError:
        print("No se tiene permiso para borrar el archivo " + ficheroXLSX)
        
    #borramos el fichero csv temporal
    try:
        os.remove(ficheroCSV)
    except FileNotFoundError:
        print("El archivo " + ficheroCSV + " no se encontró")
    except PermissionError:
        print("No se tiene permiso para borrar el archivo " + ficheroCSV)
        
    #borramos el fichero png temporal
    try:
        os.remove(ficheroPNG)
    except FileNotFoundError:
        print("El archivo " + ficheroPNG + " no se encontró")
    except PermissionError:
        print("No se tiene permiso para borrar el archivo " + ficheroPNG)

#Endpoint que es llamado por la aplicacion web para devolver el grafo global en formato png
@app.post("/api/globalgraph/{diametroF}/{diametroK}/{multiplicador_algoritmo}/{multiplicador_pesos}/{tamano_texto_nodos}/{opacidadNodos}/{habilidad}")
async def get_global_graph(diametroF: int, diametroK: int, multiplicador_algoritmo: float, multiplicador_pesos: float, tamano_texto_nodos: int, opacidadNodos: float, habilidad: bool, file: UploadFile = File(...)):
    #Guardamos el archivo excel en el servidor
    with open(file.filename, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    #definimos el nombre de los ficheros sin la extension
    nombre_ficheros = file.filename.split('.')[0]
    
    #Llamamos a la funcion que genera el grafico con el diametro de los nodos RA, el diametro de nos nodos K y los multiplicadores del algoritmo y de los pesos
    resultado = crear_datos_grafo_global(nombre_ficheros, diametroF, diametroK, multiplicador_algoritmo, multiplicador_pesos, tamano_texto_nodos, opacidadNodos, habilidad)
    #si se ha creado el grafico correctamente lo devolvemos en formato base64 y si no, devolvemos falso
    if(resultado):
        #Definimos la ubicación del archivo PNG
        file_path = nombre_ficheros + ".png"

        #Leemos la imagen y la convertimos en base64
        with open(file_path, 'rb') as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
            
        #llamamos a la funcion que borra los ficheros temporales usados para generar el grafico
        borrar_ficheros_temporales(nombre_ficheros + ".xlsx", nombre_ficheros + ".csv", nombre_ficheros + ".png")

        #Devolvemos la imagen codificada en base64 como respuesta
        return encoded_string
    else:
        return False

#Endpoint que es llamado por la aplicacion web para devolver el grafo ODS en formato png
@app.post("/api/odsgraph/{tamano_texto_nodos}/{opacidadNodos}")
async def get_ods_graph(tamano_texto_nodos: int, opacidadNodos: float, file: UploadFile = File(...)):
    #Guardamos el archivo excel en el servidor
    with open(file.filename, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    #definimos el nombre de los ficheros sin la extension
    nombre_ficheros = file.filename.split('.')[0]
    
    #Llamamos a la funcion que genera el grafico con el diametro de los nodos RA, el diametro de nos nodos K y los multiplicadores del algoritmo y de los pesos
    resultado = crear_datos_grafo_ods(tamano_texto_nodos, opacidadNodos, nombre_ficheros)

    #si se ha creado el grafico correctamente lo devolvemos en formato base64 y si no, devolvemos falso
    if(resultado):
        #Definimos la ubicación del archivo PNG
        file_path = nombre_ficheros + ".png"

        #Leemos la imagen y la convertimos en base64
        with open(file_path, 'rb') as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
            
        #llamamos a la funcion que borra los ficheros temporales usados para generar el grafico
        borrar_ficheros_temporales(nombre_ficheros + ".xlsx", nombre_ficheros + ".csv", nombre_ficheros + ".png")
        
        #Devolvemos la imagen codificada en base64 como respuesta
        return encoded_string
    else:
        return False

#Endpoint que es llamado por la aplicacion web para devolver el grafo STEAM en formato png
@app.post("/api/steamgraph/{tamano_texto_nodos}/{opacidadNodos}")
async def get_steam_graph(tamano_texto_nodos: int, opacidadNodos: float, file: UploadFile = File(...)):
    #Guardamos el archivo excel en el servidor
    with open(file.filename, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    #definimos el nombre de los ficheros sin la extension
    nombre_ficheros = file.filename.split('.')[0]
    
    #Llamamos a la funcion que genera el grafico con el diametro de los nodos RA, el diametro de nos nodos K y los multiplicadores del algoritmo y de los pesos
    resultado = crear_datos_grafo_steam(tamano_texto_nodos, opacidadNodos, nombre_ficheros)

    #si se ha creado el grafico correctamente lo devolvemos en formato base64 y si no, devolvemos falso
    if(resultado):
        #Definimos la ubicación del archivo PNG
        file_path = nombre_ficheros + ".png"

        #Leemos la imagen y la convertimos en base64
        with open(file_path, 'rb') as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
            
        #llamamos a la funcion que borra los ficheros temporales usados para generar el grafico
        borrar_ficheros_temporales(nombre_ficheros + ".xlsx", nombre_ficheros + ".csv", nombre_ficheros + ".png")
        
        #Devolvemos la imagen codificada en base64 como respuesta
        return encoded_string
    else:
        return False
    
#Endpoint para averiguar si la API esta arrancada
@app.get("/api/status")
async def get_status():
    return "OK"
      
#si el nombre de la funcion principal es 'main', entonces arrancamos la aplicacion con uvicorn en el puerto especificado
#if __name__ == "__main__":
    uvicorn.run("graph_api:app", host="0.0.0.0", port=8000, reload=True)
